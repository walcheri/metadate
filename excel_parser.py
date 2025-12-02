from openpyxl import load_workbook
from typing import List, Dict
from .base_parser import BaseParser

class ExcelParser(BaseParser):
    
    @staticmethod
    def supported_formats() -> List[str]:
        return [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel'
        ]
        
    def extract_metadata(self, file_path: str) -> List[Dict]:
        metadata = []
        
        try:
            workbook = load_workbook(file_path)
            props = workbook.properties
            
            properties = {
                'title': props.title,
                'subject': props.subject,
                'author': props.creator,
                'last_modified_by': props.lastModifiedBy,
                'created': props.created,
                'modified': props.modified,
                'keywords': props.keywords,
                'category': props.category,
                'description': props.description,
                'revision': props.revision,
                'version': props.version,
            }
            
            for key, value in properties.items():
                if value:
                    metadata.append(self._format_metadata(key, value))
                    
            # Информация о листах
            metadata.append(self._format_metadata('Sheets', len(workbook.sheetnames)))
            metadata.append(self._format_metadata('Sheet Names', ', '.join(workbook.sheetnames)))
            
        except Exception as e:
            print(f"Error reading Excel metadata: {e}")
            
        return metadata
